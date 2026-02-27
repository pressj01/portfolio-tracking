import pandas as pd
from datetime import date
from config import get_connection, EXCEL_PATH, SHEET_NAME

# Maps Excel column names (xlsm "All Accounts" sheet) to SQL column names
COLUMN_MAP = {
    "Ticker": "ticker",
    "Description": "description",
    "Type": "classification_type",
    "Price Paid": "price_paid",
    "Current Price": "current_price",
    "% Change": "percent_change",
    "Qty": "quantity",
    "Purchase Value": "purchase_value",
    "Current Value": "current_value",
    "Gain/Loss": "gain_or_loss",
    "Gain/Loss %": "gain_or_loss_percentage",
    "Div Freq": "div_frequency",
    "DRIP": "reinvest",
    "Ex-Div Date": "ex_div_date",
    "Div/Share": "div",
    "Div Paid": "dividend_paid",
    "Est. Annual Pmt": "estim_payment_per_year",
    "Monthly Income": "approx_monthly_income",
    "8% Annual Wdraw": "withdraw_8pct_cost_annually",
    "8% Monthly Wdraw": "withdraw_8pct_per_month",
    "Cash Not Reinvest": "cash_not_reinvested",
    "Cash Reinvested": "total_cash_reinvested",
    "Yield On Cost": "annual_yield_on_cost",
    "Current Yield": "current_annual_yield",
    "% of Account": "percent_of_account",
    "Shares from Div": "shares_bought_from_dividend",
    "Shares/Year": "shares_bought_in_year",
    "Shares/Month": "shares_in_month",
    "YTD Divs": "ytd_divs",
    "Total Divs Received": "total_divs_received",
    "Paid For Itself": "paid_for_itself",
}

SQL_COLUMNS = list(COLUMN_MAP.values()) + ["import_date"]


def ensure_tables_exist(conn):
    """Create all tables if they don't exist, and add new columns to existing tables."""
    cursor = conn.cursor()

    # Drop old Account table if it exists
    cursor.execute("""
        IF OBJECT_ID('dbo.Account', 'U') IS NOT NULL
            DROP TABLE dbo.Account
    """)

    # ── profiles table ────────────────────────────────────────────────────────
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLES
                       WHERE TABLE_SCHEMA='dbo' AND TABLE_NAME='profiles')
        CREATE TABLE dbo.profiles (
            id         INT IDENTITY(1,1) PRIMARY KEY,
            name       NVARCHAR(100) NOT NULL,
            created_at DATETIME DEFAULT GETDATE()
        )
    """)
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM dbo.profiles WHERE id = 1)
            INSERT INTO dbo.profiles (name) VALUES ('Owner')
    """)

    # all_account_info — drop and recreate to apply full schema (new columns added)
    # We recreate the table fresh so the schema is always correct.
    # Data is always re-imported from Excel so no data loss.
    cursor.execute("""
        IF OBJECT_ID('dbo.all_account_info', 'U') IS NULL
        CREATE TABLE dbo.all_account_info (
            ticker NVARCHAR(20) NOT NULL,
            profile_id INT NOT NULL DEFAULT 1,
            description NVARCHAR(200),
            classification_type NVARCHAR(20),
            price_paid FLOAT,
            current_price FLOAT,
            percent_change FLOAT,
            quantity FLOAT,
            purchase_value FLOAT,
            current_value FLOAT,
            gain_or_loss FLOAT,
            gain_or_loss_percentage FLOAT,
            div_frequency NVARCHAR(10),
            reinvest NVARCHAR(5),
            ex_div_date NVARCHAR(30),
            div FLOAT,
            dividend_paid FLOAT,
            estim_payment_per_year FLOAT,
            approx_monthly_income FLOAT,
            withdraw_8pct_cost_annually FLOAT,
            withdraw_8pct_per_month FLOAT,
            cash_not_reinvested FLOAT,
            total_cash_reinvested FLOAT,
            annual_yield_on_cost FLOAT,
            current_annual_yield FLOAT,
            percent_of_account FLOAT,
            shares_bought_from_dividend FLOAT,
            shares_bought_in_year FLOAT,
            shares_in_month FLOAT,
            ytd_divs FLOAT,
            total_divs_received FLOAT,
            paid_for_itself FLOAT,
            -- Legacy pillar columns (may be NULL if not in source sheet)
            hedged_anchor FLOAT,
            anchor FLOAT,
            gold_silver FLOAT,
            booster FLOAT,
            juicer FLOAT,
            bdc FLOAT,
            growth FLOAT,
            account_yield_on_cost FLOAT,
            current_yield_of_account FLOAT,
            dollars_per_hour FLOAT,
            import_date DATE,
            CONSTRAINT UQ_aai_ticker_profile UNIQUE (ticker, profile_id)
        )
    """)

    # ── profile_id migration for existing all_account_info ────────────────────
    cursor.execute("""
        IF NOT EXISTS (
            SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA='dbo' AND TABLE_NAME='all_account_info'
              AND COLUMN_NAME='profile_id'
        ) BEGIN
            DECLARE @pk_aai NVARCHAR(256);
            SELECT @pk_aai = CONSTRAINT_NAME
            FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS
            WHERE TABLE_NAME='all_account_info' AND CONSTRAINT_TYPE='PRIMARY KEY';
            IF @pk_aai IS NOT NULL
                EXEC('ALTER TABLE dbo.all_account_info DROP CONSTRAINT [' + @pk_aai + ']');
            ALTER TABLE dbo.all_account_info ADD profile_id INT NOT NULL DEFAULT 1;
        END
    """)
    cursor.execute("""
        IF NOT EXISTS (
            SELECT 1 FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS
            WHERE TABLE_NAME='all_account_info' AND CONSTRAINT_NAME='UQ_aai_ticker_profile'
        )
        ALTER TABLE dbo.all_account_info
            ADD CONSTRAINT UQ_aai_ticker_profile UNIQUE (ticker, profile_id)
    """)

    # Add new columns to existing all_account_info tables that predate this schema
    for col, coltype in [
        ("ytd_divs", "FLOAT"),
        ("total_divs_received", "FLOAT"),
        ("paid_for_itself", "FLOAT"),
    ]:
        cursor.execute(f"""
            IF NOT EXISTS (
                SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA='dbo' AND TABLE_NAME='all_account_info' AND COLUMN_NAME='{col}'
            )
            ALTER TABLE dbo.all_account_info ADD {col} {coltype}
        """)

    # holdings
    cursor.execute("""
        IF OBJECT_ID('dbo.holdings', 'U') IS NULL
        CREATE TABLE dbo.holdings (
            ticker NVARCHAR(20) NOT NULL PRIMARY KEY,
            description NVARCHAR(200),
            classification_type NVARCHAR(20),
            quantity FLOAT,
            price_paid FLOAT,
            current_price FLOAT,
            purchase_value FLOAT,
            current_value FLOAT,
            gain_or_loss FLOAT,
            gain_or_loss_percentage FLOAT,
            percent_change FLOAT
        )
    """)

    # dividends
    cursor.execute("""
        IF OBJECT_ID('dbo.dividends', 'U') IS NULL
        CREATE TABLE dbo.dividends (
            ticker NVARCHAR(20) NOT NULL PRIMARY KEY,
            div_frequency NVARCHAR(10),
            reinvest NVARCHAR(5),
            ex_div_date NVARCHAR(30),
            div_per_share FLOAT,
            dividend_paid FLOAT,
            estim_payment_per_year FLOAT,
            approx_monthly_income FLOAT,
            annual_yield_on_cost FLOAT,
            current_annual_yield FLOAT,
            ytd_divs FLOAT,
            total_divs_received FLOAT,
            paid_for_itself FLOAT
        )
    """)

    # Add new columns to existing dividends tables
    for col, coltype in [
        ("ytd_divs", "FLOAT"),
        ("total_divs_received", "FLOAT"),
        ("paid_for_itself", "FLOAT"),
    ]:
        cursor.execute(f"""
            IF NOT EXISTS (
                SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA='dbo' AND TABLE_NAME='dividends' AND COLUMN_NAME='{col}'
            )
            ALTER TABLE dbo.dividends ADD {col} {coltype}
        """)

    # income_tracking
    cursor.execute("""
        IF OBJECT_ID('dbo.income_tracking', 'U') IS NULL
        CREATE TABLE dbo.income_tracking (
            id INT IDENTITY(1,1) PRIMARY KEY,
            ticker NVARCHAR(20),
            import_date DATE,
            dividend_paid FLOAT,
            approx_monthly_income FLOAT,
            estim_payment_per_year FLOAT,
            dollars_per_hour FLOAT,
            ytd_divs FLOAT,
            total_divs_received FLOAT,
            profile_id INT NOT NULL DEFAULT 1
        )
    """)
    cursor.execute("""
        IF NOT EXISTS (
            SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA='dbo' AND TABLE_NAME='income_tracking'
              AND COLUMN_NAME='profile_id'
        )
        ALTER TABLE dbo.income_tracking ADD profile_id INT NOT NULL DEFAULT 1
    """)

    for col, coltype in [
        ("ytd_divs", "FLOAT"),
        ("total_divs_received", "FLOAT"),
    ]:
        cursor.execute(f"""
            IF NOT EXISTS (
                SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA='dbo' AND TABLE_NAME='income_tracking' AND COLUMN_NAME='{col}'
            )
            ALTER TABLE dbo.income_tracking ADD {col} {coltype}
        """)

    # pillar_weights
    cursor.execute("""
        IF OBJECT_ID('dbo.pillar_weights', 'U') IS NULL
        CREATE TABLE dbo.pillar_weights (
            ticker NVARCHAR(20) NOT NULL PRIMARY KEY,
            hedged_anchor FLOAT,
            anchor FLOAT,
            gold_silver FLOAT,
            booster FLOAT,
            juicer FLOAT,
            bdc FLOAT,
            growth FLOAT
        )
    """)

    # weekly_payouts
    cursor.execute("""
        IF OBJECT_ID('dbo.weekly_payouts', 'U') IS NULL
        CREATE TABLE dbo.weekly_payouts (
            id INT IDENTITY(1,1) PRIMARY KEY,
            pay_date DATE NOT NULL,
            week_of_month INT,
            amount FLOAT,
            profile_id INT NOT NULL DEFAULT 1,
            CONSTRAINT uq_weekly_pay_date_profile UNIQUE (pay_date, profile_id)
        )
    """)
    # Migrate existing weekly_payouts: add profile_id + update unique constraint
    cursor.execute("""
        IF NOT EXISTS (
            SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA='dbo' AND TABLE_NAME='weekly_payouts'
              AND COLUMN_NAME='profile_id'
        ) BEGIN
            IF EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS
                       WHERE TABLE_NAME='weekly_payouts' AND CONSTRAINT_NAME='uq_weekly_pay_date')
                ALTER TABLE dbo.weekly_payouts DROP CONSTRAINT uq_weekly_pay_date;
            ALTER TABLE dbo.weekly_payouts ADD profile_id INT NOT NULL DEFAULT 1;
        END
    """)
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS
                       WHERE TABLE_NAME='weekly_payouts'
                         AND CONSTRAINT_NAME='uq_weekly_pay_date_profile')
        ALTER TABLE dbo.weekly_payouts
            ADD CONSTRAINT uq_weekly_pay_date_profile UNIQUE (pay_date, profile_id)
    """)

    # monthly_payouts
    cursor.execute("""
        IF OBJECT_ID('dbo.monthly_payouts', 'U') IS NULL
        CREATE TABLE dbo.monthly_payouts (
            id INT IDENTITY(1,1) PRIMARY KEY,
            year INT NOT NULL,
            month INT NOT NULL,
            amount FLOAT,
            profile_id INT NOT NULL DEFAULT 1,
            CONSTRAINT uq_monthly_year_month_profile UNIQUE (year, month, profile_id)
        )
    """)
    # Migrate existing monthly_payouts
    cursor.execute("""
        IF NOT EXISTS (
            SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA='dbo' AND TABLE_NAME='monthly_payouts'
              AND COLUMN_NAME='profile_id'
        ) BEGIN
            IF EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS
                       WHERE TABLE_NAME='monthly_payouts' AND CONSTRAINT_NAME='uq_monthly_year_month')
                ALTER TABLE dbo.monthly_payouts DROP CONSTRAINT uq_monthly_year_month;
            ALTER TABLE dbo.monthly_payouts ADD profile_id INT NOT NULL DEFAULT 1;
        END
    """)
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS
                       WHERE TABLE_NAME='monthly_payouts'
                         AND CONSTRAINT_NAME='uq_monthly_year_month_profile')
        ALTER TABLE dbo.monthly_payouts
            ADD CONSTRAINT uq_monthly_year_month_profile UNIQUE (year, month, profile_id)
    """)

    # weekly_payout_tickers
    cursor.execute("""
        IF OBJECT_ID('dbo.weekly_payout_tickers', 'U') IS NULL
        CREATE TABLE dbo.weekly_payout_tickers (
            id INT IDENTITY(1,1) PRIMARY KEY,
            ticker NVARCHAR(20) NOT NULL,
            shares FLOAT,
            distribution FLOAT,
            total_dividend FLOAT,
            profile_id INT NOT NULL DEFAULT 1,
            CONSTRAINT uq_weekly_ticker_profile UNIQUE (ticker, profile_id)
        )
    """)
    cursor.execute("""
        IF NOT EXISTS (
            SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA='dbo' AND TABLE_NAME='weekly_payout_tickers'
              AND COLUMN_NAME='profile_id'
        ) BEGIN
            IF EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS
                       WHERE TABLE_NAME='weekly_payout_tickers' AND CONSTRAINT_NAME='uq_weekly_ticker')
                ALTER TABLE dbo.weekly_payout_tickers DROP CONSTRAINT uq_weekly_ticker;
            ALTER TABLE dbo.weekly_payout_tickers ADD profile_id INT NOT NULL DEFAULT 1;
        END
    """)
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS
                       WHERE TABLE_NAME='weekly_payout_tickers'
                         AND CONSTRAINT_NAME='uq_weekly_ticker_profile')
        ALTER TABLE dbo.weekly_payout_tickers
            ADD CONSTRAINT uq_weekly_ticker_profile UNIQUE (ticker, profile_id)
    """)

    # monthly_payout_tickers
    cursor.execute("""
        IF OBJECT_ID('dbo.monthly_payout_tickers', 'U') IS NULL
        CREATE TABLE dbo.monthly_payout_tickers (
            id INT IDENTITY(1,1) PRIMARY KEY,
            ticker NVARCHAR(20) NOT NULL,
            pay_month INT NOT NULL,
            profile_id INT NOT NULL DEFAULT 1,
            CONSTRAINT uq_monthly_ticker_month_profile UNIQUE (ticker, pay_month, profile_id)
        )
    """)
    cursor.execute("""
        IF NOT EXISTS (
            SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA='dbo' AND TABLE_NAME='monthly_payout_tickers'
              AND COLUMN_NAME='profile_id'
        ) BEGIN
            IF EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS
                       WHERE TABLE_NAME='monthly_payout_tickers'
                         AND CONSTRAINT_NAME='uq_monthly_ticker_month')
                ALTER TABLE dbo.monthly_payout_tickers DROP CONSTRAINT uq_monthly_ticker_month;
            ALTER TABLE dbo.monthly_payout_tickers ADD profile_id INT NOT NULL DEFAULT 1;
        END
    """)
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS
                       WHERE TABLE_NAME='monthly_payout_tickers'
                         AND CONSTRAINT_NAME='uq_monthly_ticker_month_profile')
        ALTER TABLE dbo.monthly_payout_tickers
            ADD CONSTRAINT uq_monthly_ticker_month_profile UNIQUE (ticker, pay_month, profile_id)
    """)

    # nav_erosion_portfolio_list — persists the user's screener ETF list between sessions
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLES
                       WHERE TABLE_NAME = 'nav_erosion_portfolio_list')
        CREATE TABLE dbo.nav_erosion_portfolio_list (
            id           INT IDENTITY(1,1) PRIMARY KEY,
            ticker       NVARCHAR(20)  NOT NULL,
            amount       FLOAT         NOT NULL,
            reinvest_pct FLOAT         NOT NULL DEFAULT 0,
            sort_order   INT           NOT NULL DEFAULT 0
        )
    """)

    # nav_erosion_saved_backtests — named saved backtest snapshots
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLES
                       WHERE TABLE_NAME = 'nav_erosion_saved_backtests')
        CREATE TABLE dbo.nav_erosion_saved_backtests (
            id          INT IDENTITY(1,1) PRIMARY KEY,
            name        NVARCHAR(200) NOT NULL,
            created_at  DATETIME      NOT NULL DEFAULT GETDATE(),
            start_date  VARCHAR(10),
            end_date    VARCHAR(10),
            rows_json   NVARCHAR(MAX) NOT NULL
        )
    """)

    # portfolio_income_sim_list — current working ETF list for the Income Simulator
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLES
                       WHERE TABLE_NAME = 'portfolio_income_sim_list')
        CREATE TABLE dbo.portfolio_income_sim_list (
            id             INT IDENTITY(1,1) PRIMARY KEY,
            ticker         NVARCHAR(20) NOT NULL,
            amount         FLOAT        NOT NULL,
            reinvest_pct   FLOAT        NOT NULL DEFAULT 0,
            yield_override FLOAT        NULL,
            sort_order     INT          NOT NULL DEFAULT 0
        )
    """)
    # Add yield_override if the table was created before this column existed
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                       WHERE TABLE_SCHEMA='dbo' AND TABLE_NAME='portfolio_income_sim_list'
                         AND COLUMN_NAME='yield_override')
        ALTER TABLE dbo.portfolio_income_sim_list ADD yield_override FLOAT NULL
    """)

    # portfolio_income_sim_saved — named saved simulations
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLES
                       WHERE TABLE_NAME = 'portfolio_income_sim_saved')
        CREATE TABLE dbo.portfolio_income_sim_saved (
            id              INT IDENTITY(1,1) PRIMARY KEY,
            name            NVARCHAR(200) NOT NULL,
            created_at      DATETIME      NOT NULL DEFAULT GETDATE(),
            mode            VARCHAR(10),
            start_date      VARCHAR(10),
            end_date        VARCHAR(10),
            market_type     VARCHAR(10),
            duration_months INT,
            rows_json       NVARCHAR(MAX) NOT NULL
        )
    """)
    # Add columns that may be missing if the table was created before they were introduced
    for col, coltype in [
        ("mode",            "VARCHAR(10)"),
        ("start_date",      "VARCHAR(10)"),
        ("end_date",        "VARCHAR(10)"),
        ("market_type",     "VARCHAR(10)"),
        ("duration_months", "INT"),
    ]:
        cursor.execute(f"""
            IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                           WHERE TABLE_SCHEMA='dbo' AND TABLE_NAME='portfolio_income_sim_saved'
                             AND COLUMN_NAME='{col}')
            ALTER TABLE dbo.portfolio_income_sim_saved ADD {col} {coltype}
        """)

    # watchlist_watching — tickers the user is interested in buying
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLES
                       WHERE TABLE_NAME = 'watchlist_watching')
        CREATE TABLE dbo.watchlist_watching (
            id          INT IDENTITY(1,1) PRIMARY KEY,
            ticker      NVARCHAR(20)  NOT NULL,
            notes       NVARCHAR(500) NULL,
            added_date  DATE          NOT NULL DEFAULT CAST(GETDATE() AS DATE),
            sort_order  INT           NOT NULL DEFAULT 0,
            CONSTRAINT uq_watchlist_watching_ticker UNIQUE (ticker)
        )
    """)

    # watchlist_sold — positions the user has exited (no unique on ticker — may sell same ticker multiple times)
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLES
                       WHERE TABLE_NAME = 'watchlist_sold')
        CREATE TABLE dbo.watchlist_sold (
            id              INT IDENTITY(1,1) PRIMARY KEY,
            ticker          NVARCHAR(20)  NOT NULL,
            buy_price       FLOAT         NULL,
            sell_price      FLOAT         NULL,
            shares_sold     FLOAT         NULL,
            sell_date       DATE          NULL,
            divs_received   FLOAT         NULL,
            notes           NVARCHAR(500) NULL,
            added_date      DATE          NOT NULL DEFAULT CAST(GETDATE() AS DATE),
            sort_order      INT           NOT NULL DEFAULT 0
        )
    """)

    # swap_candidates — user-saved swap candidate tickers per profile
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLES
                       WHERE TABLE_NAME = 'swap_candidates')
        CREATE TABLE dbo.swap_candidates (
            id         INT IDENTITY(1,1) PRIMARY KEY,
            profile_id INT          NOT NULL DEFAULT 1,
            ticker     NVARCHAR(20) NOT NULL,
            added_at   DATETIME     NOT NULL DEFAULT GETDATE(),
            CONSTRAINT uq_swap_candidates_profile_ticker UNIQUE (profile_id, ticker)
        )
    """)

    # builder_portfolios — named hypothetical portfolios per profile
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLES
                       WHERE TABLE_NAME = 'builder_portfolios')
        CREATE TABLE dbo.builder_portfolios (
            id          INT IDENTITY(1,1) PRIMARY KEY,
            profile_id  INT           NOT NULL DEFAULT 1,
            name        NVARCHAR(100) NOT NULL,
            notes       NVARCHAR(500) NULL,
            created_at  DATETIME      NOT NULL DEFAULT GETDATE(),
            updated_at  DATETIME      NOT NULL DEFAULT GETDATE(),
            CONSTRAINT uq_builder_portfolios_profile_name UNIQUE (profile_id, name)
        )
    """)

    # builder_holdings — individual positions in a builder portfolio
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLES
                       WHERE TABLE_NAME = 'builder_holdings')
        CREATE TABLE dbo.builder_holdings (
            id            INT IDENTITY(1,1) PRIMARY KEY,
            portfolio_id  INT          NOT NULL,
            ticker        NVARCHAR(20) NOT NULL,
            dollar_amount FLOAT        NOT NULL DEFAULT 0,
            added_at      DATETIME     NOT NULL DEFAULT GETDATE(),
            CONSTRAINT uq_builder_holdings_portfolio_ticker UNIQUE (portfolio_id, ticker),
            CONSTRAINT fk_builder_holdings_portfolio
                FOREIGN KEY (portfolio_id) REFERENCES dbo.builder_portfolios(id)
        )
    """)

    conn.commit()


def import_from_excel(profile_id=1):
    """Read Excel file and import into all_account_info. Returns (row_count, message)."""
    import re
    df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME, engine="openpyxl")

    # ── Detect "[Month] Income" columns and capture monthly totals ──────────
    _month_pattern = re.compile(r'^([A-Za-z]+)\s+Income$')
    _valid_ticker  = re.compile(r'^[A-Z][A-Z0-9]{0,8}$')
    _current_year  = date.today().year
    _monthly_income_updates = []  # list of (year, month_num, amount)

    for col_name in df.columns:
        if not isinstance(col_name, str):        # skip NaN / numeric column headers
            continue
        _m = _month_pattern.match(col_name)
        if not _m:
            continue
        try:
            _month_num = pd.to_datetime(_m.group(1), format='%B').month
        except Exception:
            continue
        # Sum only valid ticker rows (excludes TOTALS and summary rows)
        _mask  = df['Ticker'].apply(
            lambda t: isinstance(t, str) and t.strip() != 'TOTALS' and bool(_valid_ticker.match(t.strip()))
        )
        _total = pd.to_numeric(df.loc[_mask, col_name], errors='coerce').fillna(0).sum()
        if _total > 0:
            _monthly_income_updates.append((_current_year, _month_num, round(float(_total), 2)))

    # Keep only mapped columns that actually exist in this sheet
    excel_cols = [c for c in COLUMN_MAP.keys() if c in df.columns]
    df = df[excel_cols].copy()

    # Rename to SQL column names
    df.rename(columns=COLUMN_MAP, inplace=True)

    # Drop rows where ticker is null or blank
    df = df.dropna(subset=["ticker"])
    df = df[df["ticker"].astype(str).str.strip() != ""]

    # Keep only valid stock tickers — must be ALL UPPERCASE letters/digits (1-9 chars).
    # This filters out summary rows like TOTALS, PORTFOLIO SUMMARY, and pillar-name rows
    # such as "Anchor", "Hedged Anchor" (title-case, not real tickers).
    import re
    valid_ticker = re.compile(r'^[A-Z][A-Z0-9]{0,8}$')
    df = df[df["ticker"].astype(str).str.strip().apply(lambda t: bool(valid_ticker.match(t)))]

    # Coerce all float SQL columns to numeric — replaces text sentinel values (e.g. '--') with NaN
    float_sql_cols = [
        "price_paid", "current_price", "percent_change", "quantity",
        "purchase_value", "current_value", "gain_or_loss", "gain_or_loss_percentage",
        "div", "dividend_paid", "estim_payment_per_year", "approx_monthly_income",
        "withdraw_8pct_cost_annually", "withdraw_8pct_per_month",
        "cash_not_reinvested", "total_cash_reinvested",
        "annual_yield_on_cost", "current_annual_yield", "percent_of_account",
        "shares_bought_from_dividend", "shares_bought_in_year", "shares_in_month",
        "ytd_divs", "total_divs_received", "paid_for_itself",
    ]
    for col in float_sql_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Drop rows where current_price is NaN or zero — these are summary/header rows, not real holdings
    if "current_price" in df.columns:
        df = df[df["current_price"].notna()]

    # Add import_date and profile_id
    df["import_date"] = date.today()
    df["profile_id"] = profile_id

    conn = get_connection()
    ensure_tables_exist(conn)
    cursor = conn.cursor()

    # Clear existing data for this profile only (preserves other profiles)
    cursor.execute("DELETE FROM dbo.all_account_info WHERE profile_id = ?", profile_id)

    # Build insert using only columns that exist in the dataframe + import_date + profile_id
    cols_to_insert = [c for c in SQL_COLUMNS if c in df.columns] + ["profile_id"]
    placeholders = ", ".join(["?"] * len(cols_to_insert))
    insert_sql = (
        f"INSERT INTO dbo.all_account_info ({', '.join(cols_to_insert)}) "
        f"VALUES ({placeholders})"
    )

    row_count = 0
    for _, row in df.iterrows():
        values = []
        for col in cols_to_insert:
            val = row.get(col)
            if pd.isna(val):
                values.append(None)
            else:
                values.append(val)
        cursor.execute(insert_sql, values)
        row_count += 1

    conn.commit()

    # ── Upsert monthly income totals into monthly_payouts ────────────────────
    for _yr, _mo, _amt in _monthly_income_updates:
        cursor.execute(
            "SELECT 1 FROM dbo.monthly_payouts WHERE year = ? AND month = ? AND profile_id = ?",
            _yr, _mo, profile_id,
        )
        if cursor.fetchone() is None:
            cursor.execute(
                "INSERT INTO dbo.monthly_payouts (year, month, amount, profile_id) VALUES (?, ?, ?, ?)",
                _yr, _mo, _amt, profile_id,
            )
        else:
            cursor.execute(
                "UPDATE dbo.monthly_payouts SET amount = ? WHERE year = ? AND month = ? AND profile_id = ?",
                _amt, _yr, _mo, profile_id,
            )
    conn.commit()
    conn.close()
    return row_count, f"Successfully imported {row_count} rows from {EXCEL_PATH}."


def import_weekly_payouts(profile_id=1):
    """Import weekly payout data from Weekly_Payers sheet."""
    import openpyxl
    from datetime import datetime

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True, keep_vba=True)
    ws = wb["Weekly_Payers"]

    # Per-ticker detail from top section (rows 2-20)
    ticker_rows = []
    for row in ws.iter_rows(min_row=2, max_row=20):
        ticker = row[0].value
        if ticker is None or not isinstance(ticker, str):
            continue
        shares = row[1].value
        dist = row[8].value
        total_div = row[9].value
        ticker_rows.append((ticker, shares, dist, total_div))

    # Aggregate weekly totals from bottom section (rows 22+)
    weekly_rows = []
    for row in ws.iter_rows(min_row=22):
        pay_date_val = row[0].value
        week_val = row[1].value
        amount_val = row[2].value

        if not isinstance(pay_date_val, datetime):
            continue
        if amount_val is None:
            continue

        weekly_rows.append((
            pay_date_val.date(),
            int(week_val) if week_val is not None else None,
            float(amount_val),
        ))
    wb.close()

    conn = get_connection()
    ensure_tables_exist(conn)
    cursor = conn.cursor()

    # Clear and reimport per-ticker detail for this profile (changes each week)
    cursor.execute("DELETE FROM dbo.weekly_payout_tickers WHERE profile_id = ?", profile_id)
    for ticker, shares, dist, total_div in ticker_rows:
        cursor.execute(
            "INSERT INTO dbo.weekly_payout_tickers (ticker, shares, distribution, total_dividend, profile_id) "
            "VALUES (?, ?, ?, ?, ?)",
            ticker,
            float(shares) if shares is not None else None,
            float(dist) if dist is not None else None,
            float(total_div) if total_div is not None else None,
            profile_id,
        )

    # Insert weekly totals (skip existing dates for this profile)
    inserted = 0
    for pay_date, week, amount in weekly_rows:
        cursor.execute(
            "SELECT 1 FROM dbo.weekly_payouts WHERE pay_date = ? AND profile_id = ?",
            pay_date, profile_id,
        )
        if cursor.fetchone() is None:
            cursor.execute(
                "INSERT INTO dbo.weekly_payouts (pay_date, week_of_month, amount, profile_id) VALUES (?, ?, ?, ?)",
                pay_date, week, amount, profile_id,
            )
            inserted += 1

    conn.commit()
    conn.close()
    return inserted, f"Weekly payouts: {inserted} new rows imported, {len(ticker_rows)} tickers updated."


def import_monthly_payouts(profile_id=1):
    """Import monthly payout data from Monthly Tracking sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True, keep_vba=True)
    ws = wb["Monthly Tracking"]

    rows_to_insert = []
    for row in ws.iter_rows(min_row=2):
        year_val = row[0].value
        if year_val is None or not isinstance(year_val, (int, float)):
            continue
        year = int(year_val)
        for month_idx in range(1, 13):
            amount = row[month_idx].value
            if amount is not None:
                rows_to_insert.append((year, month_idx, float(amount)))
    wb.close()

    conn = get_connection()
    ensure_tables_exist(conn)
    cursor = conn.cursor()

    inserted = 0
    for year, month, amount in rows_to_insert:
        cursor.execute(
            "SELECT 1 FROM dbo.monthly_payouts WHERE year = ? AND month = ? AND profile_id = ?",
            year, month, profile_id,
        )
        if cursor.fetchone() is None:
            cursor.execute(
                "INSERT INTO dbo.monthly_payouts (year, month, amount, profile_id) VALUES (?, ?, ?, ?)",
                year, month, amount, profile_id,
            )
            inserted += 1

    conn.commit()
    conn.close()
    return inserted, f"Monthly payouts: {inserted} new rows imported."


def import_monthly_payout_tickers(profile_id=1):
    """Import ticker-to-month mapping from DivMonths sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True, keep_vba=True)
    ws = wb["DivMonths"]

    tickers = [cell.value for cell in ws[1] if cell.value is not None]

    ticker_months = {t: set() for t in tickers}
    for row in ws.iter_rows(min_row=2):
        for i, ticker in enumerate(tickers):
            val = row[i].value
            if val is not None and isinstance(val, (int, float)):
                ticker_months[ticker].add(int(val))
    wb.close()

    conn = get_connection()
    ensure_tables_exist(conn)
    cursor = conn.cursor()

    cursor.execute("DELETE FROM dbo.monthly_payout_tickers WHERE profile_id = ?", profile_id)
    inserted = 0
    for ticker, months in ticker_months.items():
        for month in sorted(months):
            cursor.execute(
                "INSERT INTO dbo.monthly_payout_tickers (ticker, pay_month, profile_id) VALUES (?, ?, ?)",
                ticker, month, profile_id,
            )
            inserted += 1

    conn.commit()
    conn.close()
    return inserted, f"Monthly ticker mappings: {inserted} rows imported ({len(tickers)} tickers)."


# ── Upload column name mapping (flexible, case-insensitive) ───────────────────
UPLOAD_COL_MAP = {
    'ticker':        ['ticker', 'symbol', 'stock', 'etf'],
    'quantity':      ['shares', 'quantity', 'qty', 'units', 'amount'],
    'price_paid':    ['price paid', 'cost basis', 'avg cost', 'buy price', 'price_paid'],
    'div':           ['div/share', 'dividend', 'div', 'distribution'],
    'div_frequency': ['frequency', 'div frequency', 'div freq', 'div_frequency'],
    'ex_div_date':   ['ex-div date', 'ex div date', 'ex_div_date', 'exdivdate'],
    'reinvest':      ['drip', 'reinvest', 'reinvestment'],
}


def _normalize_upload_columns(df):
    """Rename upload DataFrame columns to canonical SQL names using UPLOAD_COL_MAP."""
    lower_map = {col.lower().strip(): col for col in df.columns}
    rename = {}
    for canonical, aliases in UPLOAD_COL_MAP.items():
        for alias in aliases:
            if alias in lower_map:
                rename[lower_map[alias]] = canonical
                break
    return df.rename(columns=rename)


def import_from_upload(df, profile_id):
    """
    Import a user-uploaded DataFrame into all_account_info for the given profile.
    df must have at minimum: ticker, quantity (or shares).
    Optional: price_paid, div, div_frequency, ex_div_date, reinvest.
    Returns (row_count, message).
    """
    import re
    import yfinance as yf
    from datetime import datetime as _dt, date as _date

    df = _normalize_upload_columns(df.copy())

    # Require ticker column
    if 'ticker' not in df.columns:
        raise ValueError("Upload file must have a 'Ticker' or 'Symbol' column.")
    if 'quantity' not in df.columns:
        raise ValueError("Upload file must have a 'Shares' or 'Quantity' column.")

    # Clean tickers
    df['ticker'] = df['ticker'].astype(str).str.strip().str.upper()
    valid_ticker = re.compile(r'^[A-Z][A-Z0-9]{0,8}$')
    df = df[df['ticker'].apply(lambda t: bool(valid_ticker.match(t)))]
    df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce')
    df = df[df['quantity'].notna() & (df['quantity'] > 0)]

    if df.empty:
        raise ValueError("No valid ticker rows found in upload.")

    freq_map     = {1: 'A', 2: 'SA', 4: 'Q', 12: 'M', 52: 'W'}
    tickers_list = df['ticker'].tolist()
    ticker_str   = ' '.join(tickers_list)

    def _freq_from_count(n):
        """Infer payment frequency code from number of annual payments."""
        if n >= 45: return 'W'
        if n >= 10: return 'M'
        if n >= 3:  return 'Q'
        if n >= 2:  return 'SA'
        return 'A'

    # Single batch download: 1-year history with prices AND dividend actions.
    # This gives us reliable last-close prices, actual dividend amounts, and
    # confirmed ex-div dates from real payment history — all in one network call.
    price_map  = {}   # ticker -> last close price
    div_map    = {}   # ticker -> last per-share dividend amount
    exdiv_map  = {}   # ticker -> last ex-div date string (MM/DD/YY)
    freq_hist  = {}   # ticker -> frequency code inferred from history
    try:
        raw = yf.download(
            ticker_str, period='1y', progress=False, auto_adjust=False, actions=True
        )
        if not raw.empty:
            def _col(name):
                """Return Series (single ticker) or DataFrame (multi-ticker) for column."""
                if isinstance(raw.columns, pd.MultiIndex):
                    return raw[name] if name in raw.columns.get_level_values(0) else None
                # Simple index = single ticker; column name is the data type (e.g. 'Close')
                return raw[name] if name in raw.columns else None

            # ── Prices ──────────────────────────────────────────────────────
            close = _col('Close')
            if close is not None:
                if isinstance(close, pd.Series):
                    s = close.dropna()
                    if len(s): price_map[tickers_list[0]] = float(s.iloc[-1])
                else:
                    for t in tickers_list:
                        if t in close.columns:
                            s = close[t].dropna()
                            if len(s): price_map[t] = float(s.iloc[-1])

            # ── Dividends ───────────────────────────────────────────────────
            divs = _col('Dividends')
            if divs is not None:
                if isinstance(divs, pd.Series):
                    d = divs[divs > 0].dropna()
                    if not d.empty:
                        t0 = tickers_list[0]
                        div_map[t0]   = float(d.iloc[-1])
                        exdiv_map[t0] = d.index[-1].strftime('%m/%d/%y')
                        freq_hist[t0] = _freq_from_count(len(d))
                else:
                    for t in tickers_list:
                        if t in divs.columns:
                            d = divs[t][divs[t] > 0].dropna()
                            if not d.empty:
                                div_map[t]   = float(d.iloc[-1])
                                exdiv_map[t] = d.index[-1].strftime('%m/%d/%y')
                                freq_hist[t] = _freq_from_count(len(d))
    except Exception:
        pass

    # Per-ticker info: only needed for description / quoteType metadata now
    info_map = {}
    for t in tickers_list:
        try:
            info_map[t] = yf.Ticker(t).info or {}
        except Exception:
            info_map[t] = {}

    enriched = []
    for _, row in df.iterrows():
        t    = row['ticker']
        info = info_map.get(t, {})

        current_price = float(
            row.get('current_price', None) or
            price_map.get(t) or
            info.get('regularMarketPrice') or
            info.get('currentPrice') or 0.0
        )

        # Per-payment div: user input → last actual payment from history → annual rate
        div = float(row.get('div', None) or div_map.get(t) or info.get('dividendRate') or 0.0)

        # Ex-div date: user input → confirmed date from dividend history → info epoch
        ex_div_raw = row.get('ex_div_date', None)
        if ex_div_raw and str(ex_div_raw).strip() not in ('', 'nan', 'None'):
            ex_div_date = str(ex_div_raw).strip()
        else:
            ex_div_date = exdiv_map.get(t)
            if not ex_div_date:
                ex_ts = info.get('exDividendDate')
                if ex_ts:
                    try:
                        ex_div_date = _dt.utcfromtimestamp(ex_ts).strftime('%m/%d/%y')
                    except Exception:
                        ex_div_date = None

        # Frequency: user input → inferred from payment count → info payout frequency
        freq_raw = row.get('div_frequency', None)
        if freq_raw and str(freq_raw).strip().upper() in ('A', 'SA', 'Q', 'M', 'W', '52'):
            div_frequency = str(freq_raw).strip().upper()
        elif t in freq_hist:
            div_frequency = freq_hist[t]
        else:
            div_frequency = freq_map.get(info.get('payoutFrequency'), 'Q')

        qty = float(row['quantity'])
        price_paid_raw = row.get('price_paid', None)
        price_paid = float(price_paid_raw) if price_paid_raw and str(price_paid_raw) not in ('', 'nan') else current_price

        current_value  = current_price * qty
        purchase_value = price_paid * qty
        gain_or_loss   = current_value - purchase_value
        estim          = div * qty
        reinvest       = str(row.get('reinvest', 'N') or 'N').strip().upper()
        if reinvest not in ('Y', 'N'):
            reinvest = 'N'

        enriched.append({
            'ticker':                  t,
            'description':             info.get('longName', t)[:200] if info.get('longName') else t,
            'classification_type':     info.get('quoteType', 'ETF')[:20],
            'price_paid':              price_paid,
            'current_price':           current_price,
            'percent_change':          (gain_or_loss / purchase_value) if purchase_value else 0,
            'quantity':                qty,
            'purchase_value':          purchase_value,
            'current_value':           current_value,
            'gain_or_loss':            gain_or_loss,
            'gain_or_loss_percentage': (gain_or_loss / purchase_value) if purchase_value else 0,
            'div_frequency':           div_frequency,
            'reinvest':                reinvest,
            'ex_div_date':             ex_div_date,
            'div':                     div,
            'dividend_paid':           None,
            'estim_payment_per_year':  estim,
            'approx_monthly_income':   estim / 12,
            'annual_yield_on_cost':    (div / price_paid) if price_paid else 0,
            'current_annual_yield':    (div / current_price) if current_price else 0,
            'percent_of_account':      None,
            'ytd_divs':                None,
            'total_divs_received':     None,
            'paid_for_itself':         None,
            'import_date':             _date.today(),
            'profile_id':              profile_id,
        })

    out = pd.DataFrame(enriched)
    insert_cols = [c for c in out.columns]
    placeholders = ', '.join(['?'] * len(insert_cols))
    insert_sql = (
        f"INSERT INTO dbo.all_account_info ({', '.join(insert_cols)}) "
        f"VALUES ({placeholders})"
    )

    conn = get_connection()
    ensure_tables_exist(conn)
    cursor = conn.cursor()

    cursor.execute("DELETE FROM dbo.all_account_info WHERE profile_id = ?", profile_id)

    row_count = 0
    for _, row in out.iterrows():
        values = [None if pd.isna(v) else v for v in row.tolist()]
        cursor.execute(insert_sql, values)
        row_count += 1

    conn.commit()
    conn.close()
    return row_count, f"Imported {row_count} holdings for profile {profile_id}."
